from openpyxl import Workbook, load_workbook

class QuanLyNhanVien:
    def __init__(self, file_path="nhanvien.xlsx"):
        self.file_path = file_path

    def tao_moi_file(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["STT", "Mã", "Tên", "Tuổi"])
        wb.save(self.file_path)

    def luu_nhan_vien(self, ma, ten, tuoi):
        try:
            wb = load_workbook(self.file_path)
        except FileNotFoundError:
            self.tao_moi_file()
            wb = load_workbook(self.file_path)

        ws = wb.active
        stt = ws.max_row
        ws.append([stt, ma, ten, tuoi])
        wb.save(self.file_path)

    def doc_nhan_vien(self):
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                print(row)
        except FileNotFoundError:
            print("File không tồn tại.")

    def sap_xep_nhan_vien_theo_tuoi(self):
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            nhan_vien_data = list(ws.iter_rows(min_row=2, values_only=True))
            nhan_vien_data.sort(key=lambda x: x[3])

            for row in ws['A2:D' + str(ws.max_row)]:
                for cell in row:
                    cell.value = None
            for i, nv in enumerate(nhan_vien_data, start=2):
                ws[f"A{i}"] = i - 1
                ws[f"B{i}"] = nv[1]
                ws[f"C{i}"] = nv[2]
                ws[f"D{i}"] = nv[3]

            wb.save(self.file_path)
        except FileNotFoundError:
            print("File không tồn tại.")

qlnv = QuanLyNhanVien()

qlnv.luu_nhan_vien("NV1", "An", 18)
qlnv.luu_nhan_vien("NV2", "Lành", 22)
qlnv.luu_nhan_vien("NV3", "Giải", 20)
qlnv.luu_nhan_vien("NV4", "Thoát", 19)
qlnv.luu_nhan_vien("NV5", "Hạnh", 25)
qlnv.luu_nhan_vien("NV6", "Phúc", 24)

qlnv.doc_nhan_vien()

qlnv.sap_xep_nhan_vien_theo_tuoi()

qlnv.doc_nhan_vien()
